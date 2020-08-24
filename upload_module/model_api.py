import json, sys
import config_1 as config
config_obj = config.Config()

class model_api:
  
    def read_excel(self, csv_file, ):
        from openpyxl import load_workbook
        from openpyxl.formula import Tokenizer
        import datetime
        f_d = {}
        wb = load_workbook(csv_file, data_only=False)
        sheets = wb.get_sheet_names()
        formula_info = {}
        order_of_sheet = []
        extra_info = {}
        for sheet in sheets:
            shh_name = sheet.strip()
            ws = wb.get_sheet_by_name(sheet)
            sheetObj = ws
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    cell    = cellObj
                    formula = cell.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        formula_info.setdefault(shh_name, {})["%s_%s"%(rowid, colidx)] = formula
        wb = load_workbook(csv_file, data_only=True)
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            shh_name = sheet.strip()
            mcell_cords = {}
            for mergedCell in ws.merged_cells.ranges:
                mcell_cords[getattr(mergedCell, 'coord')] = getattr(mergedCell, "size")
            extra_info.setdefault(shh_name, {})['merge_cells'] = mcell_cords  
            data = []
            for row in range(1, ws.max_row+1):
                tmp_d = {'row': row-1, 'cols':{}}
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    fm_key = "%s_%s"%(row, col)
                    value = cell.value
                    if isinstance(cell.value, datetime.datetime):
                        value = value.strftime('%d-%b-%y')    
                    tmp_d['cols'][col -1] = {'v': value, 'fm': formula_info.get(shh_name, {}).get(fm_key, '')}
                data.append(tmp_d)
            order_of_sheet.append(shh_name)
            f_d[shh_name] = data
        return [f_d, order_of_sheet, extra_info]
    
    def clean_name(self, txt):
         if isinstance(txt, unicode) :
            txt = txt.encode('utf-8')
         try:
            txt = txt.decode('utf-8')
         except:pass
         return txt 
    def get_formula_str(self, col, colinfo):
        return ''

    def insert_model(self, ijson):
        path = ijson['path']['filename']
        industry , project, user = ijson['industry'], ijson.get('project', ''), ijson.get('user', '')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheets_info, order_of_sheet, extra_info = self.read_excel(path)
        template_id = sql_obj.max_template_id()
        if ijson.get('template_name', ''):
             template_name = ijson['template_name']
        else:
             template_name = self.clean_name(path.split('/')[-1].split('.')[0])
        sheets = []
        rows = []
        sheet_id = 1
        for sheet in order_of_sheet:
             sheet_info =  sheets_info[sheet]
             sheet = self.clean_name(sheet)  
             sheets.append((template_id, template_name, sheet_id, sheet, user, json.dumps(extra_info.get(sheet, {}))))
             for row in sheet_info:
                  for col, colinfo in row['cols'].items():
                      fromular_str = colinfo.get('fm', '') #self.get_formula_str(col, colinfo)
                      cell_alph = self.find_alp(col)
                      rows.append((template_id, sheet_id, row['row'], col, self.clean_name(colinfo['v']), 'taxonomy', fromular_str, cell_alph))
             sheet_id = sheet_id + 1
        sql_obj.insert_template(template_id, template_name, industry , project, user)
        sql_obj.insert_sheets(sheets)
        sql_obj.insert_sheet_data(rows)
        return [{'message': 'done'}]
 
    def convert_dp_format_templates(self, templates):
        avail_template  = []
        for template in templates:
            avail_template.append({'k': template[0], 'n': template[1]}) 
        return  avail_template 
        #[{'message': 'done', 'drop_down_data': avail_template}]

    def convert_grid_format_templates(self, templates):
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}, {"k":"Template", "n":"Template Name"}], 'data':[], 'map':{}}
        for rid, template in enumerate(templates, 1):
            grid_data['data'].append({'sn': {"v": rid}, 'rid': rid, 'cid': rid, "Template": {"v": template}})
            map_key = "%s_%s"%(rid,"Template")
            grid_data['map'][map_key]  = {} 
        return [grid_data] #[{'message': 'done', 'grid_data': [grid_data]}]
 
    def read_templates(self, ijson):
        industry = ijson['industry']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        templates = sql_obj.read_templates(industry) 
        avail_template  = []
        view_types = ijson['wf']
        res = {"message": "done"}
        for view_type in view_types:   
            if view_type['mod'] == 'drop_down':
                res['drop_down_data'] = self.convert_dp_format_templates(templates)
            if view_type['mod'] == 'grid':
                res['grid_data'] = self.convert_grid_format_templates(templates)
            if view_type == 'tabs':
                res['tabs_data'] = self.convert_dp_format_templates(sheets)
        return res

    def convert_dp_format_sheets(self, sheets):
        avail_sheets  = []
        for sheet in sheets:
            avail_sheets.append({'k': sheet[0], 'n':sheet[1]}) 
        return avail_sheets #[{'message': 'done', 'drop_down_data': avail_sheets}]
        
    def convert_grid_format_sheets(self, sheets):
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}, {"k":"Sheet", "n":"Sheet Name"}], 'data':[], 'map':{}}
        for rid, sheet in enumerate(sheets, 1):
            grid_data['data'].append({'sn': {"v": rid}, 'rid': rid, 'cid': rid, "Sheet": {"v": sheet}})
            map_key = "%s_%s"%(rid,"Template")
            grid_data['map'][map_key]  = {} 
        return grid_data #[{'message': 'done', 'grid_data': [grid_data]}]
        
          
    def read_sheets(self, ijson):
        industry = ijson['industry']
        template_id = ijson['template_id']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheets = sql_obj.read_sheets(template_id, industry) 
        view_types = ijson['wf']
        res = {"message": "done"}
        for view_type in view_types:   
            if view_type == 'drop_down':
                res['drop_down_data'] = self.convert_dp_format_sheets(sheets)
            if view_type == 'grid':
                res['grid_data'] = self.convert_grid_format_sheets(sheets)
            if view_type['mod'] == 'tabs':
                res['tabs_data'] = self.convert_dp_format_sheets(sheets)
        return res 

    def find_alp(self, dec_val):
        if(dec_val<26):
            return chr(65+dec_val)
        if(dec_val > 25 and dec_val < 52):
            return chr(65)+chr(65+(dec_val-26))
        if(dec_val > 51 and dec_val < 78):
            return chr(66)+chr(65+(dec_val-52))
        if(dec_val > 77 and dec_val < 104):
            return chr(67)+chr(65+(dec_val-78))
        return "@@"

    def convert_grid_format_sheetinfo(self, sheet_info):
        fs = {}
        col_def = []
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}], 'grid_data':[], 'map':{}}
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = value
             fs[row][col]['taxo'] = taxonomy
             if col not in col_def:
                col_def.append(col) 
                grid_data['col_def'].append({'k': col, 'n': cell_alph, 'mw':150, "hcls": "xl_hdr"})          
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = {'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 rr[col] = {'v': colinfo['v']}
                 grid_data['map'][str_cid+"_"+str(col)] = {'taxo': colinfo['taxo']}
              grid_data['grid_data'].append(rr)
         
        return [grid_data]

    def convert_spreadsheet_format_sheetinfo_new(self, sheet_info):
        fs = {}
        col_def = []
        grid_data = {"merges":[], "style":[], "rows":{}}
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             grid_data['rows'].setdefault(row, {})
             if not value:
                 value = ''
             grid_data['rows'][row].setdefault('cells', {}).setdefault(col, {})['text'] = value
        return grid_data

    def convert_spreadsheet_format_sheetinfo(self, sheet_info):
        fs = {}
        col_def = []
        grid_data = {'col_def':[], 'data':[], 'map':{}, "style":{}, 'nested_headers':{}}
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = value
             #if fromular_str:
             #   fs[row][col]['v'] = fromular_str
             fs[row][col]['taxo'] = taxonomy
             fs[row][col]['al'] = cell_alph
             if col not in col_def:
                col_def.append(col) 
                grid_data['col_def'].append({'width':100, 'title': cell_alph, 'type':"text"})          
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = [] #{'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 rr.append(colinfo['v']) #[col] = {'v': colinfo['v']}
                 grid_data['map'][colinfo['al']] = {'taxo': colinfo['taxo']}
              grid_data['data'].append(rr)
        grid_data['map'] =  {"Restated":{'AK5':5, "AK6": 6, "BB6":10}}     
        return grid_data

    def get_extra_info(self, template_id, sheet_id, sql_obj):
        extra_info = sql_obj.read_sheet_extra_info(template_id, sheet_id)
        return json.loads(extra_info[0])  
        
    def convert_table_format_sheetinfo(self, sheet_info):
        fs = {}
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}], 'table_data':[], 'map':{}}
        col_def = []
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = value
             fs[row][col]['taxo'] = taxonomy
             fs[row][col]['cell_alph'] = cell_alph
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = {'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 if col not in col_def:
                    col_def.append(col) 
                    grid_data['col_def'].append({'k': col, 'n': colinfo["cell_alph"]})          
                 rr[col] = {'v': colinfo['v']}
                 grid_data['map'][str_cid+"_"+str(col)] = {'taxo': colinfo['taxo']}
              grid_data['table_data'].append(rr)
        return grid_data
 
    def convert_merge_cells(self, mc_lst):
        import re
        res = {}
        for cord, rc_info in mc_lst.items():  
            start, end = cord.split(':')
            res[start] = [rc_info.get('columns', 1), rc_info.get('rows', 1)] 
        return res

    def sheet_gridinfo_all(self, ijson):
        template_id = ijson['template_id']
        #sheet_id   = ijson['sheet_id']
        sheets  = ijson.get('sheets')
        industry = ijson.get('industry', '')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        if not sheets:
            sheets = map(lambda x: x[0], self.read_sheets(template_id, industry))
        res = {"message": "done"}
        view_types = ijson['wf']
        for sheet_info in sheets:
            sheet_id = sheet_info
            sheet_info = sql_obj.read_sheet_data(template_id, sheet_id) 
            extra_info = self.get_extra_info(template_id, sheet_id, sql_obj)
            for view_type in view_types:   
                if view_type['mod'] == 'spreadsheet':
                    res[sheet_id] = {}
                    res[sheet_id]['spreadsheet_data'] = self.convert_spreadsheet_format_sheetinfo(sheet_info)
                    if extra_info.get('merge_cells', {}):
                        res['sheet_id']['spreadsheet_data']['merge_cells'] = self.convert_merge_cells(extra_info['merge_cells'])
        return res 

    def sheet_gridinfo(self, ijson):
        template_id = ijson['template_id']
        sheet_id   = ijson['sheet_id']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheet_info = sql_obj.read_sheet_data(template_id, sheet_id) 
        extra_info = self.get_extra_info(template_id, sheet_id, sql_obj)
        view_types = ijson['wf']
        res = {"message": "done"}
        for view_type in view_types:   
            if view_type['mod'] == 'grid':
                res['grid_data'] = self.convert_grid_format_sheetinfo(sheet_info)
            if view_type['mod'] == 'table':
                res['table_data'] = self.convert_table_format_sheetinfo(sheet_info)
            if view_type['mod'] == 'spreadsheet':
                res['spreadsheet_data'] = self.convert_spreadsheet_format_sheetinfo(sheet_info)
                if extra_info.get('merge_cells', {}):
                    res['spreadsheet_data']['merge_cells'] = self.convert_merge_cells(extra_info['merge_cells'])
            if view_type['mod'] == 'spreadsheet_new':
                res['spreadsheet_data'] = self.convert_spreadsheet_format_sheetinfo_new(sheet_info)
                if extra_info.get('merge_cells', {}):
                    res['spreadsheet_data']['merges'] = extra_info['merge_cells'].keys() 
        return res 
        
        


if __name__ == '__main__': 
    obj = model_api()
    #ijson = {"path":{"filename": "/var/www/html/muthu/test_excel/GA_Merge.xlsx"}, "user":"user", "project":"INC", "industry": "2", "template_id": 1,'sheet_id':2}
    ijson = {"cgi":"Read Sheetinfo","inp_data":[{"v":9,"k":"sheet_id"},{"v":2,"k":"industry"},{"v":"","k":"template_name"},{"v":1,"k":"template_id"}],"wf":[{"cid":"tas_cd_5","mod":"spreadsheet_new","data_cond":{"data load":True,"data select":False,"highlight":False},"Serv":"Read Sheetinfo"}], "sheets":[8,9], "industry": 2, "template_id": 1, "sheet_id":2}
    #print json.dumps(obj.insert_model(ijson))
    #print json.dumps(obj.read_templates(ijson))
    #print json.dumps(obj.read_sheets(ijson))
    print json.dumps(obj.sheet_gridinfo(ijson))
    #print json.dumps(obj.sheet_gridinfo_all(ijson))
