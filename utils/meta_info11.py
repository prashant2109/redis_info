import json, hashlib, ast, sets, binascii, copy, os, shelve, sys, datetime, sqlite3
import db.get_conn as get_conn
conn_obj    = get_conn.DB()
import html_entity_to_single_char
en_obj = html_entity_to_single_char.html_entity_to_single_char()
from collections import defaultdict as DD, OrderedDict as OD
class MetaInfo(object):
    def __init__(self, pypath="/root/databuilder_train_ui/tenkTraining/Data_Builder_Training_Copy/pysrc"):
        self.config = json.load(open(pypath+'/config.json'))
        self.grm_mrks   = [
                            '&#8217;',
                            "'",
                            '"',
                            ]
        self.alpha      = {'a':1, 'b':1, 'c':1, 'd':1, 'e':1, 'f':1, 'g':1, 'h':1, 'i':1, 'j':1, 'k':1, 'l':1, 'm':1, 'n':1, 'o':1, 'p':1, 'q':1, 'r':1, 's':1, 't':1 ,'u':1, 'v':1, 'w':1, 'x':1, 'y':1, 'z':1,' ':1}
        self.number  = {'1': 1, '0': 1, '3': 1, '2': 1, '5': 1, '4': 1, '7': 1, '6': 1, '9': 1, '8': 1}
        self.roman_number  = {'i': 1, 'ii':1, 'iii':1, 'iv':1,'v':1,'vi':1, 'vii':1, 'viii':1,'ix':1, 'x':1}
        pass

    def clean_xmls(self, xml):
        xml     = xml.replace('@@@', ':@:') 
        fxml    = []
        for x in xml.split(':@:'):
            xml_lst     = []
            for tx in x.split('#'):
                if tx.strip():
                    xml_lst.append(tx)
            x   = str('#'.join(xml_lst))
            fxml.append(x)
        return ':@:'.join(fxml)

    def remove_grm_mrks(self, txt):
        for grm in self.grm_mrks:
            txt = txt.replace(grm, '')
        txt = ' '.join(txt.split())
        return txt
    def convert_html_entity(self, txt):
        if isinstance(txt, unicode):
            txt = txt.encode('utf-8')
        txt = en_obj.convert(txt).replace('\\\\\\\\xe2\\\\\\\\x80\\\\\\\\x93', '-').replace('\\\\xe2\\\\x80\\\\x93', '-').replace('\\xe2\\x80\\x93', '-').replace('\\\\\\\\xe2\\\\\\\\x80\\\\\\\\x99', "'").replace('\\\\xe2\\\\x80\\\\x99', "'").replace('\\xe2\\x80\\x99', "'").replace('\xc2\xa0', '') #.replace('\xc3\xb3', 'o')
        if isinstance(txt, unicode):
            txt = txt.encode('utf-8')
        return txt

    def get_company_info(self, ijson):
        db          = self.config['cinfo'] #'/root/eMB_db/company_info/compnay_info.db'
        
        conn, cur   = conn_obj.sqlite_connection(db)
        cids        = ijson.get('cids', [])
        cids        = map(lambda x:str(x), cids)
        if ijson.get('industry_type'):
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency from company_info  where industry_type = '%s'"%(ijson['industry_type'])
        elif cids:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency from company_info  where toc_company_id in (%s)"%(', '.join(cids))
        else:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency from company_info"
        cur.execute(sql)
        res         = cur.fetchall()
        cur.close()
        conn.close()
        tmpd    = {}
        machine_id  = '122'
        for rr in res:
            row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency = rr
            toc_company_id          = str(toc_company_id)
            #model_number    = 'TEST'
            tmpd[int(toc_company_id)]    = {'company_name': company_name ,'model_number': model_number ,'project_id': project_id,'deal_id': toc_company_id ,'template_name': type_of_company,'industry_type': industry_type,'reporting':reporting_year,'filing':filing_frequency, 'org_company_name':company_display_name, 'machine_id':machine_id}
            tmpd['%s_%s'%(project_id, toc_company_id)]    = {'company_name': company_name ,'model_number': model_number ,'project_id': project_id,'deal_id': toc_company_id ,'template_name': type_of_company,'industry_type': industry_type,'reporting':reporting_year,'filing':filing_frequency, 'org_company_name':company_display_name, 'machine_id':machine_id}
        if ijson.get('ret', '') == 'Y':
            return tmpd
        f_ar    = []
        ks  = tmpd.keys()
        ks.sort()
        for k in ks:
            f_ar.append(tmpd[k])
        res = [{'message':'done', 'data':f_ar}]
        return res
    def get_company_info_cid(self, ijson):
        db          = self.config['cinfo'] #'/root/eMB_db/company_info/compnay_info.db'
        
        conn, cur   = conn_obj.sqlite_connection(db)
        cids        = ijson.get('cids', [])
        cids        = map(lambda x:str(x), cids)
        if ijson.get('industry_type'):
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency, user_industry_type, user_filing_frequency, user_reporting_year from company_info  where industry_type = '%s'"%(ijson['industry_type'])
        elif cids:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency, user_industry_type, user_filing_frequency, user_reporting_year from company_info  where toc_company_id in (%s)"%(', '.join(cids))
        else:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency, user_industry_type, user_filing_frequency, user_reporting_year from company_info"
        cur.execute(sql)
        res         = cur.fetchall()
        cur.close()
        conn.close()
        tmpd    = {}
        machine_id  = '122'
        cid_tup = []
        for rr in res:
            row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency, user_industry_type, user_filing_frequency, user_reporting_year = rr
            toc_company_id          = str(toc_company_id)
            cid = '_'.join(map(str, [project_id, toc_company_id]))
            c_tup = map(int, [project_id, toc_company_id])
            cid_tup.append(c_tup) 
            #model_number    = 'TEST'
            tmpd[cid]    = {'company_name': company_name ,'model_number': model_number ,'project_id': project_id,'deal_id': toc_company_id ,'template_name': type_of_company,'industry_type': industry_type,'reporting':reporting_year,'filing':filing_frequency, 'org_company_name':company_display_name, 'machine_id':machine_id}
        if ijson.get('ret', '') == 'Y':
            return tmpd
        f_ar    = []
        cid_tup.sort()
        #ks  = tmpd.keys()
        #ks.sort()
        ks = map(lambda x:'_'.join(map(str, x)), cid_tup)
        for k in ks:
            f_ar.append(tmpd[k])
        res = [{'message':'done', 'data':f_ar}]
        return res

    def read_company_info(self, ijson):
        ijson['ret']    = 'Y'
        return self.get_company_info(ijson)


    def get_company_infon(self, ijson):
        #db          = self.config['cinfo'] #
        db = '/mnt/eMB_db/company_info/4433_compnay_info.db'
        
        conn, cur   = conn_obj.sqlite_connection(db)
        cids        = ijson.get('cids', [])
        cids        = map(lambda x:str(x), cids)
        if cids:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency from company_info  where toc_company_id in (%s)"%(', '.join(cids))
        else:
            sql         = "select row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency from company_info"
        cur.execute(sql)
        res         = cur.fetchall()
        cur.close()
        conn.close()
        tmpd    = {}
        machine_id  = '122'
        for rr in res:
            row_id,company_name,company_display_name,project_id,toc_company_id,type_of_company,model_number,industry_type,internal_status,external_status,reporting_year,filing_frequency = rr
            toc_company_id          = str(toc_company_id)
            #model_number    = 'TEST'
            tmpd[int(toc_company_id)]    = {'company_name': company_name ,'model_number': model_number ,'project_id': project_id,'deal_id': toc_company_id ,'template_name': type_of_company,'industry_type': industry_type,'reporting':reporting_year,'filing':filing_frequency, 'org_company_name':company_display_name, 'machine_id':machine_id}
        if ijson.get('ret', '') == 'Y':
            return tmpd
        f_ar    = []
        ks  = tmpd.keys()
        ks.sort()
        for k in ks:
            f_ar.append(tmpd[k])
        res = [{'message':'done', 'data':f_ar}]
        return res

    def read_company_infon(self, ijson):
        ijson['ret']    = 'Y'
        return self.get_company_infon(ijson)



    def get_quid(self, text):
        m = hashlib.md5()
        m.update(text)
        quid = m.hexdigest()
        return quid

    def get_bbox_frm_xml(self, txn1, table_id, t_x, ret_f=None):
        valbox      = []
        x_c         = ''
        done_d      = {}
        for x in [t_x]:
           bbox = txn1.get(str(table_id)+':$$:'+str(x))
           if not bbox:
               bbox = txn1.get(str(table_id)+':$$:'+self.get_quid(str(x)))
           #print str(table_id)+':$$:'+str(x), txn1.get(str(table_id)+':$$:'+str(x))
           if bbox:
                bbox1 = ast.literal_eval(bbox)
                if not bbox1:continue
                    #print bbox1
                cnt_flg = 0
                for tmpbbox in bbox1[0]:
                    if not tmpbbox or tuple(tmpbbox) in done_d:continue
                    done_d[tuple(tmpbbox)]  = 1
                    valbox.append(tmpbbox)
                    cnt_flg = 1

        if not valbox:
            for  cxml in t_x.split(':@:'):
                for x in cxml.split('#'):
                   if x and x_c =='':
                        x_c= x 
                   
                   bbox = txn1.get(str(table_id)+':$$:'+str(x))
                   if not bbox:
                       bbox = txn1.get(str(table_id)+':$$:'+self.get_quid(str(x)))
                   #print str(table_id)+':$$:'+str(x), txn1.get(str(table_id)+':$$:'+str(x))
                   if bbox:
                        bbox1 = ast.literal_eval(bbox)
                        #print bbox1
                        if not bbox1 or tuple(bbox1[0][0]) in done_d:continue
                        done_d[tuple(bbox1[0][0])]  = 1
                        valbox.append(bbox1[0][0])
                        continue
                   #print str(table_id)+':$$:'+str(x), txn1.get(str(table_id)+':$$:'+str(x.split('@')[0]))
                   bbox = txn1.get(str(table_id)+':$$:'+str(x.split('@')[0]))
                   if not bbox:
                       bbox = txn1.get(str(table_id)+':$$:'+self.get_quid(str(x.split('@')[0])))
                   if bbox:
                        bbox1 = ast.literal_eval(bbox)
                        #print bbox1
                        if tuple(bbox1[0][0]) in done_d:continue
                        done_d[tuple(bbox1[0][0])]  = 1
                        valbox.append(bbox1[0][0])
        if ret_f == 'Y':
            if x_c == '':
                x_c = 'x'+table_id+'_'+table_id
            return valbox, x_c
        return valbox

    def get_db_path(self, ijson):
        company_name    = ijson['company_name']
        mnumber         = ijson['model_number']
        dtime           = ijson.get('dtime', '')
        model_number    = mnumber
        db_file         = '/mnt/eMB_db/%s/%s/mt_data_builder.db'%(company_name, model_number)
        if ijson.get('taxo_flg', '') == 1:
            if dtime:
                db_file         = '/mnt/eMB_db/%s/%s/taxo_data_builder_%s.db'%(company_name, model_number, dtime)
            else:
                db_file         = '/mnt/eMB_db/%s/%s/taxo_data_builder.db'%(company_name, model_number)
        return db_file

    def get_cyear(self, lines):
        doc_years   = {}
        for line in lines[1:]:
            line    = line.split('\t')
            if len(line) < 8:continue
            line    = map(lambda x:x.strip(), line)
            ph      = line[3]+line[7]
            try:
                if 'FY' == line[3] or 'Q4' == line[3]:
                    year    = int(ph[-4:])
                    doc_years[year] = 1
            except:continue
        c_years      = doc_years.keys()
        c_years.sort(reverse=True)
        if not c_years:
            yr = datetime.datetime.today().year
            c_years.extend([yr])
        return c_years[0]
        

    def sheet_id_map(self):
        from collections import defaultdict as dd
        db_file     = '/mnt/eMB_db/node_mapping.db'
        conn, cur   = conn_obj.sqlite_connection(db_file)
        sql   = "select main_header, sub_header, sheet_id, node_name, description from node_mapping where review_flg = 0"
        cur.execute(sql)
        tres    = cur.fetchall()
        #print rr, len(tres)
        ddict   = dd(set)
        for tr in tres:
            #print tr
            main_header, sub_header, sheet_id, node_name, description = map(str, tr)
            ddict[sheet_id] = [main_header, sub_header, node_name, description]
        return ddict

    def get_all_docids(self, company_name, model_number):
        db_file         = '/mnt/eMB_db/%s/%s/tas_company.db'%(company_name, model_number)
        conn, cur       = conn_obj.sqlite_connection(db_file)
        sql             = "select doc_id from company_meta_info"
        cur.execute(sql)
        res             = cur.fetchall()
        doc_d   = {}
        for rr in res:
            doc_d[str(rr[0])]   = 1
        return doc_d

    def get_all_tablessm(self, company_name, model_number):
        db_file     = os.path.join('/mnt/eMB_db/', company_name, model_number, 'tas_company.db')   
        conn, cur       = conn_obj.sqlite_connection(db_file)
        stmt = "select doc_id, page_no, row_id from table_dict_phcsv_info"
        cur.execute(stmt)
        res = cur.fetchall()
        norm_res_list = []
        tdict = {}
        for row in res:
            doc_id, page_no, table_id = map(str, row[:])
            norm_res_list.append((doc_id, page_no, table_id))
            tdict[table_id] = 1
        return tdict

    def custom_sheet_id_map(self):
        from collections import defaultdict as dd
        db_file     = '/mnt/eMB_db/node_mapping.db'
        conn, cur   = conn_obj.sqlite_connection(db_file)
        sql   = "select main_header, sub_header, sheet_id, node_name, description from custom_node_mapping where review_flg = 0"
        cur.execute(sql)
        tres    = cur.fetchall()
        #print rr, len(tres)
        ddict   = dd(set)
        for tr in tres:
            #print tr
            main_header, sub_header, sheet_id, node_name, description = map(str, tr)
            ddict[sheet_id] = [main_header, sub_header, node_name, description]
        return ddict
        
    def custom_info(self, conn, cur):
        sheet_id_map = self.custom_sheet_id_map()
        try:
            sql_cstm             = "select sheet_id, doc_id, doc_name, table_id from custom_table_group_mapping"
            cur.execute(sql_cstm)
            res_cstm             = cur.fetchall()
        except:res_cstm = []
        print 'sss', res_cstm
        
        tmpres = []
        table_type_m = {}
        for rr in res_cstm:
            sheet_id, doc_id, doc_name, table_id    = rr
            if str(doc_id) not in doc_d:continue
            main_header, sub_header, node_name, description = sheet_id_map.get(str(sheet_id), ['','','',''])
            #print rr, (main_header, sub_header, node_name, description)
            if f_tables and node_name not in f_tables:continue
            #print rr, main_header
            #if main_header not in ['Main Table']:continue
            desc_sp = description.split('-')
            desc_sp = map(lambda x:x.strip(), desc_sp[:])
            desc_sp = list(sets.Set(desc_sp))
            if len(desc_sp) == 1:
               description = desc_sp[0]
            table_type_m[node_name]    = description
            tmpres.append((node_name, doc_id, table_id, main_header))
        return table_type_m, tmpres 

    def get_main_table_info_new(self, company_name, model_number, f_tables=[]):
        tdict = {}
        if str(model_number) == '50':
            tdict = self.get_all_tablessm(company_name, model_number)
        #elif str(model_number) == '60':
        #    tdict = self.get_all_tablessm(company_name, model_number)
        doc_d           = self.get_all_docids(company_name, model_number)
        sheet_id_map = self.sheet_id_map()
        #print 'SSSS', sheet_id_map
        m_tables        = {}
        doc_m_d         = {}
        rev_m_tables    = {}
        db_file         = '/mnt/eMB_db/%s/%s/tas_company.db'%(company_name, model_number)
        conn, cur       = conn_obj.sqlite_connection(db_file)
        sql             = "select sheet_id, doc_id, doc_name, table_id from table_group_mapping"
        cur.execute(sql)
        res             = cur.fetchall()
        tmpres          = []
        table_type_m    = {}
        '''
        try:
            t1_dct, i_lst = self.custom_info(conn, cur)
        except:t1_dct, i_lst = {}, []
        '''
        
        for rr in res:
            sheet_id, doc_id, doc_name, table_id    = rr
            if str(doc_id) not in doc_d:continue
            main_header, sub_header, node_name, description = sheet_id_map.get(str(sheet_id), ['','','',''])
            #print rr, (main_header, sub_header, node_name, description)
            if f_tables and node_name not in f_tables:continue
            #print rr, main_header
            #if main_header not in ['Main Table']:continue
            desc_sp = description.split('-')
            desc_sp = map(lambda x:x.strip(), desc_sp[:])
            desc_sp = list(sets.Set(desc_sp))
            if len(desc_sp) == 1:
               description = desc_sp[0]
            table_type_m[node_name]    = description
            tmpres.append((node_name, doc_id, table_id, main_header))
        cur.close()
        conn.close()
        m_tables        = {}
        doc_m_d         = {}
        rev_m_tables    = {}
        main_headers    = {}
        for rr in tmpres:
            table_type, doc_id, table_id_str, main_header    = rr
            doc_id      = str(doc_id)
            table_type  = str(table_type)
            for table_id in table_id_str.split('^!!^'):
                if not table_id:continue
                if tdict and table_id not in tdict:continue
                main_headers[main_header]    = 1
                table_id            = str(table_id)
                m_tables[table_id]  = table_type
                rev_m_tables.setdefault(table_type, {})[table_id]   = 1
                doc_m_d[table_id]   = doc_id
        table_type_m['main_header'] = main_headers.keys()
        return m_tables, rev_m_tables, doc_m_d, table_type_m

    def get_tt_data_60(self, company_name, model_number):
        db_path = '/mnt/eMB_db/%s/%s/tas_company.db'%(company_name, model_number)
        conn = sqlite3.connect(db_path) 
        cur  = conn.cursor()
        read_qry = 'SELECT doc_id, table_id, table_type FROM table_dict_phcsv_info;'
        cur.execute(read_qry)
        table_data = cur.fetchall()
        conn.close()
        m_tables, rev_m_tables, doc_m_d, table_type_m = {}, {}, {}, {}        
        for row in table_data:
            doc_id, table_id, table_type = map(str, row)
            m_tables[table_id] = table_type
            rev_m_tables.setdefault(table_type, {})[table_id] = 1
            doc_m_d[table_id] = doc_id
        table_type_m['main_header'] = rev_m_tables.keys()
        return m_tables, rev_m_tables, doc_m_d, table_type_m

    def get_main_table_info(self, company_name, model_number, f_tables=[]):
        return self.get_main_table_info_new(company_name, model_number, f_tables)
        #elif model_number == '60':
        #    return self.get_tt_data_60(company_name, model_number)
        db_file         = '/mnt/eMB_db/%s/%s/tas_company.db'%(company_name, model_number)
        conn, cur       = conn_obj.sqlite_connection(db_file)
        if f_tables:
            sql             = "select table_type, doc_id, table_id from table_map where table_type in (%s)"%(','.join(map(lambda x:'"'+x+'"', f_tables)))
        else:
            sql             = "select table_type, doc_id, table_id from table_map"
        cur.execute(sql)
        res             = cur.fetchall()
        cur.close()
        conn.close()
        m_tables        = {}
        doc_m_d         = {}
        rev_m_tables    = {}
        table_type_m    = {}
        for rr in res:
            table_type, doc_id, table_id_str    = rr
            doc_id      = str(doc_id)
            table_type  = str(table_type)
            table_type_m[table_type]    = table_type
            for table_id in table_id_str.split(':^:'):
                if not table_id:continue
                table_id            = str(table_id).strip()
                m_tables[table_id]  = table_type
                rev_m_tables.setdefault(table_type, {})[table_id]   = 1
                doc_m_d[table_id]   = doc_id
        if rev_m_tables:
            table_type_m['main_header'] = rev_m_tables.keys()[0]
        else:
            table_type_m['main_header'] = ''
        return m_tables, rev_m_tables, doc_m_d, table_type_m

    def gen_taxonomy(self, txt):
        alpha      = {'a':1, 'b':1, 'c':1, 'd':1, 'e':1, 'f':1, 'g':1, 'h':1, 'i':1, 'j':1, 'k':1, 'l':1, 'm':1, 'n':1, 'o':1, 'p':1, 'q':1, 'r':1, 's':1, 't':1 ,'u':1, 'v':1, 'w':1, 'x':1, 'y':1, 'z':1}
        tmp_ar  = []
        for c in txt:
            if c.lower() in alpha:
                tmp_ar.append(c)
        return ''.join(tmp_ar)


    def numToWords(self, num,join=True):
        '''words = {} convert an integer number into words'''
        units = ['','one','two','three','four','five','six','seven','eight','nine']
        teens = ['','eleven','twelve','thirteen','fourteen','fifteen','sixteen', \
                 'seventeen','eighteen','nineteen']
        tens = ['','ten','twenty','thirty','forty','fifty','sixty','seventy', \
                'eighty','ninety']
        thousands = ['','thousand','million','billion','trillion','quadrillion', \
                     'quintillion','sextillion','septillion','octillion', \
                     'nonillion','decillion','undecillion','duodecillion', \
                     'tredecillion','quattuordecillion','sexdecillion', \
                     'septendecillion','octodecillion','novemdecillion', \
                     'vigintillion']
        words = []
        if num==0: words.append('zero')
        else:
            numStr = '%d'%num
            numStrLen = len(numStr)
            groups = (numStrLen+2)/3
            numStr = numStr.zfill(groups*3)
            for i in range(0,groups*3,3):
                h,t,u = int(numStr[i]),int(numStr[i+1]),int(numStr[i+2])
                g = groups-(i/3+1)
                if h>=1:
                    words.append(units[h])
                    words.append('hundred')
                if t>1:
                    words.append(tens[t])
                    if u>=1: words.append(units[u])
                elif t==1:
                    if u>=1: words.append(teens[u])
                    else: words.append(tens[t])
                else:
                    if u>=1: words.append(units[u])
                if (g>=1) and ((h+t+u)>0): words.append(thousands[g]+',')
        if join: return ' '.join(words)
        return words

    def read_table_data(self, table_id, txn_m):
        table_id    = str(table_id)
        normalized_data = []
        rc_d    = {}
        for l in ['HGH', 'VGH', 'GV', 'GH']:
            ids = txn_m.get(l+'_'+table_id)
            if not ids:continue
            for c_id in ids.split('#'):
                tid, r, c   = c_id.split('_')
                r           = int(r)
                c           = int(c)
                x           = txn_m.get('XMLID_'+c_id)
                t           = binascii.a2b_hex(txn_m.get('TEXT_'+c_id))
                cs          = int(txn_m.get('colspan_'+c_id))
                rs          = int(txn_m.get('rowspan_'+c_id))
                rc_d.setdefault(r, {})[c]   = (r, c, t, l, cs, rs, l, x)
        rows    = rc_d.keys()
        rows.sort()
        for r in rows:
            cols    = rc_d[r].keys()
            cols.sort()
            normalized_data.append(map(lambda x:rc_d[r][x], cols))
        return normalized_data

    def parse_ijson(self, ijson):
        company_name    = ijson['company_name']
        mnumber         = ijson['model_number']
        model_number    = mnumber
        deal_id         = ijson['deal_id']
        project_id      = ijson['project_id']
        company_id      = "%s_%s"%(project_id, deal_id)
        return company_name, model_number, deal_id, project_id, company_id

    def gen_taxonomy(self, txt):
        alpha      = {'a':1, 'b':1, 'c':1, 'd':1, 'e':1, 'f':1, 'g':1, 'h':1, 'i':1, 'j':1, 'k':1, 'l':1, 'm':1, 'n':1, 'o':1, 'p':1, 'q':1, 'r':1, 's':1, 't':1 ,'u':1, 'v':1, 'w':1, 'x':1, 'y':1, 'z':1}
        tmp_ar  = []
        for c in txt:
            if c.lower() in alpha:
                tmp_ar.append(c)
        return ''.join(tmp_ar)

    def gen_taxonomy_alpha_num(self, txt):
        alpha      = {'a':1, 'b':1, 'c':1, 'd':1, 'e':1, 'f':1, 'g':1, 'h':1, 'i':1, 'j':1, 'k':1, 'l':1, 'm':1, 'n':1, 'o':1, 'p':1, 'q':1, 'r':1, 's':1, 't':1 ,'u':1, 'v':1, 'w':1, 'x':1, 'y':1, 'z':1, '0':1,'1':1,'2':1,'3':1,'4':1, '5':1,'6':1,'7':1, '8':1, '9':1}
        tmp_ar  = []
        for c in txt:
            if c.lower() in alpha:
                tmp_ar.append(c)
        return ''.join(tmp_ar)


    def order_by_table_structure_by_col(self, f_taxo_arr, table_ids):
        tmptable_ids   = map(lambda x:x[1], table_ids)
        table_match_d   = {}
        for ti, dd in enumerate(f_taxo_arr):
            ks      = dd['ks']
            for table_id, c_id in ks:
                table_match_d.setdefault(table_id, {}).setdefault(ti, {})[c_id] = 1
        tabl_ids    = table_match_d.keys()
        tabl_ids.sort(key=lambda x:tmptable_ids.index(x))
        #tabl_ids.sort(key=lambda x:len(table_match_d[x].keys()), reverse=True)
        final_arr   = []
        for table_id in tabl_ids:
            inds    = table_match_d[table_id].keys()
            inds.sort()
            #print '\n==========================================================='
            #print table_id, sorted(inds)
            #inds.sort(key=lambda x:int(table_match_d[table_id][x].keys()[0].split('_')[2]))
            inds.sort(key=lambda x:int(sorted(map(lambda x1:int(x1.split('_')[2]), table_match_d[table_id][x].keys()))[0]))
            #print 'Ordered ', inds
            if not final_arr:
                final_arr   = inds
            else:
                m_d = list(sets.Set(final_arr).intersection(sets.Set(inds)))
                deletion    = {}
                tmp_arr     = []
                ftmp_arr     = []
                for t in inds:
                    if t in m_d:
                        ftmp_arr    = []
                        if tmp_arr:
                            deletion[t] = copy.deepcopy(tmp_arr[:])
                            #tmp_arr = []    
                        continue
                    tmp_arr.append(t)
                    ftmp_arr.append(t)
                done_d  = {}
                m_d.sort(key=lambda x:final_arr.index(x))
                
                for t in m_d:
                    if t in deletion:
                        tmp_arr = []
                        for t1 in deletion[t]:
                            if t1 not in done_d:
                                tmp_arr.append(t1)
                                done_d[t1]  = 1
                        index       = final_arr.index(t)
                        final_arr   = final_arr[:index]+tmp_arr+final_arr[index:]    
                
                if ftmp_arr:
                    final_arr   = final_arr+ftmp_arr
            
            #print 'FINAL ', final_arr
        missing = sets.Set(range(len(f_taxo_arr))) - sets.Set(final_arr)
        if len(final_arr) > len(f_taxo_arr):
            print 'Duplicate ', final_arr
            sys.exit()
        if len(missing):
            print 'missing ', list(missing)
            sys.exit()
        f_taxo_arr  = map(lambda x:f_taxo_arr[x], final_arr)
        return f_taxo_arr

    def read_excel_data(self, csv_file, sheet_names=[], ret_flg=None):
        import xlrd
        xl_workbook = xlrd.open_workbook(csv_file)
        if not sheet_names:
            sheet_names = xl_workbook.sheet_names()
        f_d = {}
        f_cell_d = {}
        for i in range(0,len(sheet_names)):
            xl_sheet        = xl_workbook.sheet_by_name(sheet_names[i])
            num_rows    = xl_sheet.nrows
            num_cols    = xl_sheet.ncols
            herader     = []
            data        = []
            for row_idx in range(0, num_rows):
                tmp_d           = {}
                if row_idx ==0:
                    for col_idx in range(0, num_cols):
                        val = xl_sheet.cell(row_idx, col_idx).value
                        if isinstance(val, unicode):
                            val = val.encode('utf-8')
                        if isinstance(val, str):
                            val = ' '.join(val.strip().split())
                        herader.append(val)
                    continue
                #print '\n=================================='
                for col_idx in range(0, num_cols):
                    val = xl_sheet.cell(row_idx, col_idx).value
                    cell_name   = xlrd.formula.cellname(row_idx, col_idx)
                    f_cell_d.setdefault(' '.join(sheet_names[i].split()).strip(), {})[cell_name]    = (len(data), herader[col_idx])
                    if isinstance(val, unicode):
                        val = val.encode('utf-8')
                    if isinstance(val, str):
                        val = ' '.join(val.strip().split())
                    else:
                        val = str(val)
                    #print [herader[col_idx], val, xl_sheet.cell(row_idx, col_idx).value]
                    cmt     = ''
                    page    = ''
                    ph      = ''
                    tmp_d[herader[col_idx]] = val
                    if ret_flg == 'Y':
                        tmp_d[('RC', herader[col_idx])] = (row_idx, col_idx)
                data.append(tmp_d)
            f_d[' '.join(sheet_names[i].split()).strip()] = (herader, data)
        if ret_flg == 'Y':
            f_d['FINFO_TT'] = f_cell_d
        return f_d

    def read_from_shelve(self, fname, default={}):
        if os.path.isfile(fname):
            sh = shelve.open(fname, 'r')
            data = sh.get('data', default)
            sh.close()
            return data
        return default

    def read_formula(self, fname):
        formula_d   = {}
        from openpyxl import load_workbook
        from openpyxl.formula import Tokenizer
        excel_data = {} 
        data_wb = load_workbook(filename=fname, data_only=True)
        wb = load_workbook(filename=fname, data_only=False)
        visible_sheets = [idx for idx, sheet in enumerate(wb._sheets) if sheet.sheet_state == "visible"]
        sheet_names = wb.get_sheet_names()
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    formula = cellObj.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        tok = Tokenizer(formula)
                        flist   = []
                        opr_ar  = []
                        for t in tok.items:
                            if t.type == 'OPERAND':
                                if opr_ar:
                                    dd  = (t.value, opr_ar[-1], 't')
                                else:
                                    dd  = (t.value, '+', 't')
                                for op in opr_ar[:-1]:
                                    tdd  = ('', op, 'v')
                                    flist.append(tdd)
                                flist.append(dd)
                                opr_ar  = []
                            else:
                                opr_ar.append(t.value)
                        for op in opr_ar:
                            tdd  = ('', op, 'v')
                            flist.append(tdd)
                        formula_d.setdefault(' '.join(sheet_name.split()).strip(), {})[(rowid, colidx)]  = (formula, flist)
                        pass    
        return formula_d

    def get_deriv_phs(self, ijson):
        if ijson.get('project_name', ''):
            company_name    = ijson['company_name']
            mnumber         = ijson['model_number']
            model_number    = mnumber
            deal_id         = ijson['deal_id']
            project_id      = ijson['project_id']
            db_file         = self.get_db_path(ijson)
            conn, cur   = conn_obj.sqlite_connection(db_file)
            sql         = "select reporting_type, total_years, derived_ph, periods, ind_group_flg, taxo_flg, excel_config_str from company_config where project_name='%s'"%(ijson['project_name'])
            try:
                cur.execute(sql)
                res = cur.fetchall()
            except:
                res = []
            conn.close()
            deriv_phs   = {}
            for rr in res:
                reporting_type, total_years, derived_ph, periods, ind_group_flg, taxo_flg, excel_config_str  = rr
                if periods:
                    for dph in periods.split('##'):
                        if dph.strip():
                            deriv_phs[dph]  = 1
            return  deriv_phs
        return {}



    def read_ph_user_formula(self, ijson, group_id=None, tph_formula_d={}, all_tids={}):
        company_name    = ijson['company_name']
        mnumber         = ijson['model_number']
        model_number    = mnumber
        deal_id         = ijson['deal_id']
        project_id      = ijson['project_id']
        company_id      = "%s_%s"%(project_id, deal_id)
        table_type  = ijson['table_type']
        opr     = {
                    '=' : 1,
                    '+' : 1,
                    '-' : 1,
                    '*' : 1,
                    '/' : 1,
                    '(' : 1,
                    ')' : 1,
                    }
        deriv_phs   = self.get_deriv_phs(ijson)
        #print ijson
        #print deriv_phs
        #sys.exit()
        db_file         = self.get_db_path(ijson)
        conn, cur   = conn_obj.sqlite_connection(db_file)
        if group_id:
            group_id_sp = [group_id]
            if '-' in group_id:
                group_id_sp.append(group_id.split('-')[0])
            group_id_sp = map(lambda x:'"'+x+'"', group_id_sp)
            sql         = 'select row_id, ph, formula, formula_type, formula_str, group_id from ph_derivation where table_type="%s" and group_id in (%s)'%(table_type, ','.join(group_id_sp))
        else:
            sql         = 'select row_id, ph, formula, formula_type, formula_str, group_id from ph_derivation where table_type="%s"'%(table_type)
        #print db_file, sql
        try:
            cur.execute(sql)
            res         = cur.fetchall()
        except:
            res  = []

        ftype_map_d = {
                        'FORMULA'       : 'F',
                        'SFORMULA'      : 'F',
                        'CELLFORMULA'   : 'CELL F',
                        'PTYPEFORMULA'  : 'PTYPE F',
                        'USERFORMULA'   : 'USER F',
                        'DIRECTFORMULA' : 'DIRECT F',
                        'ACROSSFORMULA' : 'ACROSS F',
                        'WITHINFORMULA' : 'WITHIN F',
                        }
        ph_formula_d    = copy.deepcopy(tph_formula_d)
        for rr in res:
            row_id, ph, formula, formula_type, formula_str, tgroup_id    = rr
            if not group_id and tgroup_id:continue
            row_id  = 'RID-'+str(row_id)
            if formula_type == 'TH':
                ph_formula_d[(formula_type, ph)]    = formula #, {})[int(tid)]  = 1
            elif formula_type == 'TTYPE':
                ph_formula_d[(formula_type, ph)]    = formula #, {})[int(tid)]  = 1
            elif formula_type == 'Ignore_taxo':
                for tid in formula_str.split('##'):
                    if not tid:continue
                    ph_formula_d.setdefault((formula_type, formula), {})[int(tid)]  = 1
            elif formula_type == 'CF':
                sph                 = ph.split('-')
                F_taxoid, CF_cellid, rowflg, colflg, year_diff = formula.split('^') 
                year_diff   = int(year_diff)
                ph_formula_d.setdefault(row_id, {'i':{}, 'v':(F_taxoid, CF_cellid, sph[0], sph[1])})
                ph_formula_d[row_id]['r']  = True if rowflg == '1' else False
                ph_formula_d[row_id]['c']  = True if colflg == '1' else False
                for t_k in formula_str.split('##'):
                    c_sp    = t_k.split('$')
                    taxo_id = c_sp[0]
                    for cell_id in c_sp[1:]:
                        cell_id = cell_id.split('^^')[0]
                        if rowflg == '1':
                            ph_formula_d.setdefault(('CF', taxo_id), {})[sph[0]]   = (row_id, F_taxoid, CF_cellid, sph[1], year_diff)
                        else:
                            ph_formula_d.setdefault(('CF', taxo_id), {})[cell_id]   = (row_id, F_taxoid, CF_cellid, sph[1], year_diff)
                        ph_formula_d[row_id]['i'][(taxo_id, cell_id)] = 1
                        
                continue
            elif formula_type == 'SCALE':
                for t_k in formula_str.split('##'):
                    c_sp    = t_k.split('$')
                    taxo_id = c_sp[0]
                    for cell_id in c_sp[1:]:
                        cell_id = cell_id.split('^^')[0]
                        ph_formula_d.setdefault(('S', taxo_id), {})[cell_id]    = (row_id, formula)
                continue
            elif formula_type == 'PH' and formula and '=' in formula:
                operands    = []
                tmp_ar      = []
                tph, op_str  = formula.split('=')
                if tph not in deriv_phs:continue
                for c in op_str:
                    #print '\t',[c, c in opr]
                    if c in opr:
                        if tmp_ar:
                            operands.append({'t':'op', 'v':''.join(tmp_ar)})
                            tmp_ar  = []
                        operands.append({'t':'opr', 'v':c})
                    elif c:
                        tmp_ar.append(c)
                #print [tmp_ar]
                if tmp_ar:
                    operands.append({'t':'op', 'v':''.join(tmp_ar)})
                #print operands
                if operands:
                    for t_k in formula_str.split('##'):
                        c_sp    = t_k.split('$')
                        taxo_id = c_sp[0]
                        for cell_id in c_sp[1:]:
                            cell_id = cell_id.split('^^')[0]
                            #print [taxo_id, cell_id, formula, operands]
                            ph_formula_d.setdefault(taxo_id, {}).setdefault(cell_id, []).append((formula, operands, row_id))
                            ph_formula_d[taxo_id][('rid', cell_id)] = row_id
            elif formula_type in ftype_map_d: #['FORMULA', 'CELLFORMULA', 'PTYPEFORMULA', 'USERFORMULA']:
                #print ph_formula
                operands    = []
                done_taxo   = {}
                r_rid       = ''
                fph_year    = ''
                fph_pt      = ''
                op_d    = {}
                for topr in formula.split('$$'):
                    topr_sp = topr.split('@@')
                    tid, operator, t_type, g_id, ftype = topr_sp[:5]
                    if tid in ['', 'undefined']:
                        ftype   = 'v'
                    if g_id == 'undefined':
                        g_id    = ''
                        if group_id:
                            g_id    = str(group_id)
                    
                    if t_type == 'undefined':
                        t_type  = table_type
                    c, s,vt = '', '', ''
                    fph          = ''
                    fyear       = ''
                    to_dealid   = int(deal_id)
                    try:
                        to_dealid   =  int(topr_sp[9])
                    except:pass
                    if ijson.get('grpid') and g_id == '' and t_type == table_type and to_dealid == int(deal_id):
                        g_id = str(ijson['grpid'])
                    if operator == '=':
                        if g_id and ijson.get('grpid') and g_id != str(ijson['grpid']):
                            operands    = []
                            break
                        
                    try:
                        fph = topr_sp[8]
                        if operator == '=':
                            fph_year  = int(fph[-4:])
                            fph_pt      = topr_sp[8]
                        fyear  = int(fph[-4:])
                        fph = topr_sp[8][:-4]
                    except:pass
                    year_diff   = 0
                    if fph_year and fyear and operator != '=':
                        year_diff   = fph_year - fyear
                    
                    try:
                        c, s,vt = topr_sp[5], topr_sp[6], topr_sp[7]
                    except:pass
                    if c == 'undefined':
                        c   = ''
                    if s == 'undefined':
                        s   = ''
                    if vt == 'undefined':
                        vt   = ''
                    if ftype != 'v' and (t_type == table_type and all_tids and tid not in all_tids):
                        if ijson.get('consider_docs', []) and ijson.get('RETURN_TXT') == 'Y':
                            operands    = []
                            break
                        continue
                    if (t_type, g_id, tid) in done_taxo and (ftype != 'v') and  formula_type not in ['CELLFORMULA', 'PTYPEFORMULA']:
                        continue
                    if operator == '=':
                        r_rid   = tid
                    done_taxo[(t_type, g_id, tid)]  = 1
                    op_d[tid]    = 1
                    if formula_type in ['PTYPEFORMULA'] :
                        operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator, 'c':c, 's':s,'vt':vt, 'period':fph, 'fyear':fyear, 'yd':year_diff, 'rid':row_id, 'to_dealid':to_dealid, 'i_f_type':formula_type})
                    elif formula_type in ['CELLFORMULA']:
                        operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator, 'c':c, 's':s,'vt':vt, 'rid':row_id, 'k':fph+str(fyear), 'to_dealid':to_dealid, 'i_f_type':formula_type, 'ph':fph+str(fyear)})
                    else:
                        operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator, 'c':c, 's':s,'vt':vt, 'rid':row_id, 'to_dealid':to_dealid, 'i_f_type':formula_type})
                #print operands
                if formula_type in ['PTYPEFORMULA']  and fph_pt:
                    fph_pt  = fph_pt[:-4]
                if operands and r_rid:
                    f_taxos = done_taxo.keys()
                    f_taxos.sort()
                    for t_k in formula_str.split('##'):
                        c_sp    = t_k.split('$')
                        taxo_id = c_sp[0]
                        for cell_id in c_sp[1:]:
                            if formula_type in ['CELLFORMULA', 'PTYPEFORMULA'] :
                                #ph_formula_d.setdefault((ftype_map_d[formula_type], taxo_id), {})[cell_id]  = (row_id, operands)
                                #ph_formula_d.setdefault(('ALL_'+ftype_map_d[formula_type], taxo_id, cell_id), {})[tuple(f_taxos)]  = (row_id, operands)
                                if fph_pt:
                                    ph_formula_d.setdefault((ftype_map_d[formula_type], taxo_id), {})[fph_pt]  = (row_id, operands)
                                    ph_formula_d.setdefault(('ALL_'+ftype_map_d[formula_type], taxo_id, fph_pt), {})[tuple(f_taxos)]  = (row_id, operands)
                            else:
                                ph_formula_d.setdefault('FORM', {}).update(op_d)
                                ph_formula_d[(ftype_map_d[formula_type], taxo_id)]  = (row_id, operands)
                                ph_formula_d.setdefault(('ALL_'+ftype_map_d[formula_type], taxo_id), {})[tuple(f_taxos)]  = (row_id, operands)
            elif formula_type == 'SFORMULA':
                #print ph_formula
                operands    = []
                r_rid   = ''
                done_taxo   = {}
                for topr in formula.split('$$'):
                    topr_sp = topr.split('@@')
                    tid, operator, t_type, g_id, ftype = topr_sp[:5]
                    if all_tids and tid not in all_tids:
                        if ijson.get('consider_docs', []) and ijson.get('RETURN_TXT') == 'Y':
                            operands    = []
                            break
                        continue
                    if operator == '=':
                        r_rid   = tid
                    if tid in done_taxo:
                        continue
                    done_taxo[tid]  = 1
                    operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator})
                if operands and r_rid:
                    ph_formula_d[('SYS F', r_rid)]  = (row_id, operands)
                    f_taxos = done_taxo.keys()
                    f_taxos.sort()
                    ph_formula_d.setdefault(('ALL_SYS F', r_rid), {})[tuple(f_taxos)]  = (row_id, operands)
                    for t_k in formula_str.split('##'):
                        c_sp    = t_k.split('$')
                        taxo_id = c_sp[0]
                        for cell_id in c_sp[1:]:
                            cell_id = cell_id.split('^^')[0]
                            if cell_id.split('^^')[1:]:
                                ph_formula_d.setdefault(("OP", row_id), {})[(taxo_id, cell_id)]  = cell_id.split('^^')[1:]
                    
            elif formula_type == 'CFFORMULA' and 0:
                rph, rowflg = ph.split('^') 
                #print ph
                operands    = []
                for topr in formula.split('$$'):
                    topr_sp = topr.split('@@')
                    tid, operator, t_type, g_id, ftype = topr_sp[:5]
                    tid, phk, tph   = tid.split('^')
                    year_diff   = int(rph[-4:]) - int(tph[-4:])
                    if all_tids and tid not in all_tids:
                        if ijson.get('consider_docs', []) and ijson.get('RETURN_TXT') == 'Y':
                            operands    = []
                            break
                        continue
                    if rowflg == '1':
                        operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator, 'period':tph[:-4], 'yd':year_diff})
                    else:
                        operands.append({'txid':tid, 'type':ftype, 't_type':t_type, 'g_id':g_id, 'op':operator, 'k':phk})
                #print operands
                if operands:
                    for t_k in formula_str.split('##'):
                        c_sp    = t_k.split('$')
                        taxo_id = c_sp[0]
                        for cell_id in c_sp[1:]:
                            cell_id = cell_id.split('^^')[0]
                            if rowflg == '1':
                                ph_formula_d.setdefault(('CF_F', taxo_id), {})[rph[:-4]]  = (row_id, rowflg, operands)
                            else:
                                ph_formula_d.setdefault(('CF_F', taxo_id), {})[cell_id]  = (row_id, rowflg, operands)
        return ph_formula_d

    def remove_index_num(self, txt):
        allnum  = sets.Set('0123456789')
        txt = txt.lower()
        l   = len(txt)
        idx = 0
        for c in txt:
            if c == '.' or c == ')':
                if idx > 0 and idx < l: 
                    tmptxt  = txt[:idx]
                    if tmptxt in self.alpha:
                        return txt[idx+1:].strip()
                    if tmptxt in self.roman_number:
                        return txt[idx+1:].strip()
                    if tmptxt in self.number or sets.Set(tmptxt).issubset(allnum):
                        return txt[idx+1:].strip()
                break
            idx += 1
        return txt

    def remove_foot_node(self, txt):
        txt = txt.lower()
        txt = self.remove_index_num(txt)
        ntxt, foot_node = self.foot_node_detection(txt)
        if foot_node:
            txt    = ' '.join(ntxt.split())
        return txt

    def foot_node_detection(self, txt):
        
        if '(' not in txt or ')' not in txt:
            return txt, [] 
        i   = 0
        l   = len(txt)
        foot_nodes  = []
        tmp_gtxt     = ''
        while i < l:
            char    = txt[i]
            st_ind  = i
            if '(' == char:
                char_ar = []
                f   = 0
                tind    = 0
                f_alpha = 0
                for t_char in txt[i+1:]:
                    if t_char in self.alpha:
                        f_alpha =1
                    char_ar.append(t_char)
                    tind += 1
                    if t_char == ')':
                        f   = 1
                        break
                if f == 1:
                    tmp_txt = ''.join(''.join(char_ar[:-1]).strip().split())
                    if (not tmp_txt or len(tmp_txt) == 1) or (len(tmp_txt) > 1 and f_alpha == 0):
                        foot_nodes.append((txt[st_ind:st_ind+tind+1], (st_ind, st_ind+tind+1)))
                        st_ind  = st_ind+tind+1
                i   += tind
            tmp_gtxt += txt[st_ind:i+1]
            i+= 1
        return tmp_gtxt, foot_nodes


    def get_date(self, mrange, phn):
        month   = '01'
        year    = int( phn[-4:])
        if mrange:
            mrange  = mrange.split('-')
            if mrange[-1].strip():
                month1   = datetime.datetime.strptime(mrange[0].strip().lower(), '%b').strftime('%m')
                month2   = datetime.datetime.strptime(mrange[-1].strip().lower(), '%b').strftime('%m')
                if int(month1) < int(month2):
                    year    = year-1
                    month   = month2
        date    = '%s-%s-01'%(year, month)
                    
        return date

    def read_user_selected_col_map_info(self, ijson):
        company_name    = ijson['company_name']
        mnumber         = ijson['model_number']
        model_number    = mnumber
        deal_id         = ijson['deal_id']
        project_id      = ijson['project_id']
        company_id      = "%s_%s"%(project_id, deal_id)
        table_type      = ijson['table_type']
        group_id        = ijson.get('grpid', '')
        db_file = '/mnt/eMB_db/%s/%s/mt_data_builder.db'%(company_name, model_number)
        conn, cur       = conn_obj.sqlite_connection(db_file)
        crt_qry = """CREATE TABLE IF NOT EXISTS user_selected_gv_info(row_id INTEGER PRIMARY KEY AUTOINCREMENT, table_type TEXT, group_id TEXT, ph TEXT, col_id TEXT, taxo_map TEXT)"""
        cur.execute(crt_qry)
        read_qry = """SELECT ph, col_id, taxo_map FROM user_selected_gv_info WHERE table_type='%s' and group_id='%s'"""%(table_type, group_id)
        cur.execute(read_qry)
        table_data = cur.fetchall() 
        conn.close()
        res_dct = {}
        for row in table_data:
            ph, col_id, taxo_map = row
            col_id_lst = col_id.split('#')
            tids = {}
            for mp in taxo_map.split('~~'):
                if not mp:continue
                key, val = mp.split('#')
                tids[key] = val
            res_dct.setdefault(ph, {'phs':[], 'tids':{}})
            res_dct[ph]['phs']  += col_id_lst
            res_dct[ph]['tids'].update(tids)
        return res_dct

    def alter_table_coldef(self, conn, cur, table_name, coldef):
        col_info_stmt   = 'pragma table_info(%s);'%table_name
        cur.execute(col_info_stmt)
        all_cols        = cur.fetchall()
        cur_coldef      = set(map(lambda x:str(x[1]), all_cols))
        exists_coldef   = set(coldef)
        new_cols        = list(exists_coldef.difference(cur_coldef))
        col_list = []
        for new_col in coldef:
            if new_col not in new_cols:continue
            col_list.append(' '.join([new_col, 'TEXT']))
        for col in col_list:
            alter_stmt = 'alter table %s add column %s;'%(table_name, col)
            #print alter_stmt
            try:
                cur.execute(alter_stmt)
            except:pass
        conn.commit()
        return 'done'

    def clean_dic(self, dd):
        r_d = {}
        for k, v in dd.items():
            r_d[k]  = v
        return r_d

