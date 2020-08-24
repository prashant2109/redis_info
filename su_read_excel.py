from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import get_column_letter
import datetime, json, re, string

class read_excel:

    def get_column_index(self, col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num 


    def getValue(self, operand, sheet_info, sheet):
        sheet_name = sheet
        if "!" in operand:
            sheet_name, cell_id = operand.split('!')
        else:
            cell_id = operand
        column, row = re.split('\d+', cell_id)[0], re.split('[a-zA-Z]+', cell_id)[1]
        col_ind = self.get_column_index(column)
        cell_val = sheet_info[sheet_name][int(row)+1]['cols'][col_ind]
        return cell_val['v']


    def formula_expension(self, tok_list, sheet_info, sheet):
        result_list = []
        for each in tok_list:
            print '---', each
            if each[1] == "OPERAND" and each[2] == "RANGE":
               print 'in', each[0]
               temp_res = self.getValue(each[0], sheet_info, sheet)
               print 'out', temp_res
               result_list.append(temp_res)
        return result_list


    def formula_eval(self, excel_info, order_of_sheet, extra_info, formula_info):
        for sheet, sheet_formulas in formula_info.items():
            for rc, formula_str in sheet_formulas.items():
                print 'formula', formula_str
                tok = Tokenizer(formula_str)
                tok_list = [[t.value, t.type, t.subtype] for t in tok.items]
                ff = self.formula_expension(tok_list, excel_info, sheet)
                print [tok, ff]    

    def read_excel(self, csv_file):
        from openpyxl import load_workbook
        from openpyxl.formula import Tokenizer
        import datetime
        f_d = {}
        wb = load_workbook(csv_file, data_only=False)
        sheets = wb.get_sheet_names()
        formula_info = {}
        order_of_sheet = []
        extra_info = {}
        for sheet in sheets:
            shh_name = sheet.strip()
            ws = wb.get_sheet_by_name(sheet)
            sheetObj = ws
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    cell    = cellObj
                    formula = cell.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        formula_info.setdefault(shh_name, {})["%s_%s"%(rowid, colidx)] = formula
        wb = load_workbook(csv_file, data_only=True)
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            shh_name = sheet.strip()
            mcell_cords = {}
            for mergedCell in ws.merged_cells.ranges:
                mcell_cords[getattr(mergedCell, 'coord')] = getattr(mergedCell, "size")
            extra_info.setdefault(shh_name, {})['merge_cells'] = mcell_cords  
            data = []
            for row in range(1, ws.max_row+1):
                tmp_d = {'row': row-1, 'cols':{}}
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    fm_key = "%s_%s"%(row, col)
                    value = cell.value
                    if isinstance(cell.value, datetime.datetime):
                        value = value.strftime('%d-%b-%y')    
                    tmp_d['cols'][col -1] = {'v': value, 'fm': formula_info.get(shh_name, {}).get(fm_key, '')}
                data.append(tmp_d)
            order_of_sheet.append(shh_name)
            f_d[shh_name] = data
        self.formula_eval(f_d, order_of_sheet, extra_info, formula_info)
        return [f_d, order_of_sheet, extra_info]

if __name__ == '__main__':
    obj = read_excel() 
    res = obj.read_excel("/var/www/html/muthu/test_excel/GA_Merge.xlsx") 
    #print json.dumps(res)
