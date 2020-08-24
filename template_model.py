import json, sys, os, copy
import re
import config_1 as config
config_obj = config.Config()

class model_api:

    def read_excel_kendo(self, csv_file):
        from openpyxl import load_workbook
        from openpyxl.formula import Tokenizer
        import datetime
        f_d = []
        wb = load_workbook(csv_file, data_only=False)
        sheets = wb.get_sheet_names()
        formula_info = {}
        order_of_sheet = []
        extra_info = {}
        activeSheet = ''
        for sheet in sheets:
            shh_name = sheet.strip()
            ws = wb.get_sheet_by_name(sheet)
            if not activeSheet:
                activeSheet = sheet
            sheetObj = ws
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    cell    = cellObj
                    formula = cell.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        formula_info.setdefault(shh_name, {})["%s_%s"%(rowid, colidx)] = formula
        wb = load_workbook(csv_file, data_only=True)
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            shh_name = sheet.strip()
            mcell_cords = []
            for mergedCell in ws.merged_cells.ranges:
                mcell_cords.append(getattr(mergedCell, 'coord'))
            extra_info.setdefault(shh_name, {})['merge_cells'] = mcell_cords  
            data = []
            for row in range(1, ws.max_row+1):
                tmp_d = {'index': row-1, 'cells':[]}
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    fm_key = "%s_%s"%(row, col)
                    value = cell.value
                    if isinstance(cell.value, datetime.datetime):
                        value = value.strftime('%d-%b-%y')    
                    if not value:
                        value = ''
                    ctemp = {'value': value, 'index': col-1}
                    f_cell = formula_info.get(shh_name, {}).get(fm_key, '')
                    if f_cell:
                        ctemp['formula'] = {'src': f_cell}
                    else:
                        ctemp['value'] = value
                    if f_cell and value in ['#Value', '#DIV/0!']:
                        ctemp['value'] = {'core': value}
                    print ['ttt', f_cell, value, row, col, ctemp, f_cell]
                    tmp_d['cells'].append(ctemp)
                data.append(tmp_d)
            order_of_sheet.append(shh_name)
            f_d.append({"name": shh_name, "rows": data, "column":[], "mergedCells": mcell_cords}) #[shh_name] = data
        fs_info = {'activeSheet':activeSheet, 'sheets': f_d}
  
    def read_excel(self, csv_file):
        from openpyxl import load_workbook
        from openpyxl.formula import Tokenizer
        import datetime
        f_d = {}
        wb = load_workbook(csv_file, data_only=False)
        sheets = wb.get_sheet_names()
        formula_info = {}
        order_of_sheet = []
        extra_info = {}
        for sheet in sheets:
            shh_name = sheet.strip()
            ws = wb.get_sheet_by_name(sheet)
            sheetObj = ws
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    cell    = cellObj
                    formula = cell.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        formula_info.setdefault(shh_name, {})["%s_%s"%(rowid, colidx)] = formula
        wb = load_workbook(csv_file, data_only=True)
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            shh_name = sheet.strip()
            mcell_cords = {}
            for mergedCell in ws.merged_cells.ranges:
                mcell_cords[getattr(mergedCell, 'coord')] = getattr(mergedCell, "size")
            extra_info.setdefault(shh_name, {})['merge_cells'] = mcell_cords  
            data = []
            for row in range(1, ws.max_row+1):
                tmp_d = {'row': row-1, 'cols':{}}
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    fm_key = "%s_%s"%(row, col)
                    value = cell.value
                    if isinstance(cell.value, datetime.datetime):
                        value = value.strftime('%d-%b-%y')    
                    index_col_val = ''
                    if index_column != '':
                        index_col_val = ws.cell(row = row, column = index_column).value    
                    if row == 1 and value and (value.lower() == 'index' or value.lower() == 'level'):
                        index_column = col
                    tmp_d['cols'][col -1] = {'v': value, 'fm': formula_info.get(shh_name, {}).get(fm_key, ''), 'level': index_col_val}
                data.append(tmp_d)
            order_of_sheet.append(shh_name)
            f_d[shh_name] = data
        return [f_d, order_of_sheet, extra_info]
    
    def clean_name(self, txt):
         if isinstance(txt, unicode) :
            txt = txt.encode('utf-8')
         try:
            txt = txt.decode('utf-8')
         except:pass
         return txt 
    def get_formula_str(self, col, colinfo):
        return ''

    def wget_files(self, path):
        import time
        fname = time.strftime("%Y%m%d-%H%M%S")
        abs_path = "%s%s.xlsx"%("/var/www/html/sudhakar/", fname)
        cmd  = "wget -q -O '%s' '%s'"%(abs_path, path)
        os.system(cmd)
        return abs_path

    def convert_g(self, ijson):
        path  = self.wget_files(ijson['path']['filename'])
        return self.read_excel_kendo(path)

    def get_sheets_new(self, ijson):
        #f = open('/var/www/html/yo.txt')
        #import json
        #data = f.read()
        #f.close()
        #return json.loads(data)
        if ijson.get('link', 'N') == 'N':
            upload_path = config_obj.upload_path
            ijson['path'] = upload_path+ijson['path'] 
        print ijson['path'], ijson
        try:
            print 'socket execution started'
            import muthu_translate.socket_client_utils2 as up
            up_obj = up.socket_client_utils2('172.16.20.52', '1604') 
            data  = up_obj.send_socket(ijson)
            print 'socket execution ended', data
            return json.loads(data)
        except:
            print 'socket execution failed python insertion running'
            return self.convert_g(ijson)
        

    def insert_model(self, ijson):
        path = ijson['path']
        industry , project, user, ttype = ijson.get('industry', ''), ijson.get('project', ''), ijson.get('user', ''), ijson.get('type', '')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        if ijson.get('template_name', ''):
             template_name = ijson['template_name']
        else:
             template_name = self.clean_name(path.split('/')[-1].split('.')[0])
        only_temp = 'N'
        if not ijson.get('path', ''):
            data = {}
        else:
            only_temp = 'Y'
            data = self.get_sheets_new(ijson)
        temp_extra_info = {}
        temp_extra_info['activesheet'] = data.get('activesheet')
        if ijson.get('template_id',''):
            template_id = int(ijson['template_id'])
            sql_obj.deleted_temp_info(template_id)
        else:
            template_id = sql_obj.max_template_id()
            sql_obj.insert_template(template_id, template_name, industry , project, ttype, user, json.dumps(temp_extra_info))
        if only_temp == 'N':
            return [{'message': 'done'}]
        sheets = []
        rows = []
        sheet_id = 1
        taxo_creation = 'N'
        for sheet_info in data['sheets']:
             print 'sheete',  sheet_info['name']
             sheet = self.clean_name(sheet_info['name'])  
             if sheet == config_obj.taxonomy_sheet:
                taxo_creation = 'Y'
             extra_info = copy.deepcopy(sheet_info)
             del extra_info['rows']
             if sheet_info['name'] == data['activeSheet']:
                extra_info['activesheet'] = 1
             sheets.append((template_id, template_name, sheet_id, sheet, user, json.dumps(extra_info)))
             for row in sheet_info['rows']:
                  for colinfo in row['cells']:
                      col = colinfo['index']
                      fromular_str = colinfo.get('formula', '') #self.get_formula_str(col, colinfo)
                      cell_alph = self.find_alp(col)
                      level_id = colinfo.get('level', '')
                      rows.append((template_id, sheet_id, row['index'], col, str(colinfo), 'taxonomy', str(fromular_str), cell_alph, level_id))
             sheet_id = sheet_id + 1
        sql_obj = sqlite_api.sqlite_api(db_path)
        sql_obj.insert_sheets(sheets)
        sql_obj.insert_sheet_data(rows)
        if taxo_creation == 'Y' and ttype == 'fixed': 
            import pyapi
            py_obj = pyapi.PYAPI()
            print 'insertion code', template_id, project
            py_obj.insert_taxo_to_db(project, template_id) 
        return [{'message': 'done'}]

    def update_cell_property(self, ijson):
        template_id = ijson['template_id']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        for sheet_id, sinfo in ijson['data'].items():
            for rc, rc_val in sinfo.items():
                r,c  = map(lambda x: int(x), rc.split('_'))
                old_rc_info = sql_obj.get_row_id_info(template_id, sheet_id, r,c) 
                old_rc_info.update(rc_val)
                sql_obj.update_cell_value(json.dumps(old_rc_info), template_id, sheet_id, r,c)
        return [{'messge':'done'}]
                
            
            

    def insert_new_template(self, ijson):
        path = ijson['path']
        industry , project, user, ttype = ijson.get('industry', ''), ijson.get('project', ''), ijson.get('user', ''), ijson.get('type', '')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        #data = ijson['data'] #self.get_sheets_new(ijson)
        data = self.get_sheets_new(ijson)
        template_id = sql_obj.max_template_id()
        if ijson.get('template_name', ''):
             template_name = ijson['template_name']
        else:
             template_name = self.clean_name(path.split('/')[-1].split('.')[0])
        temp_extra_info = {}
        temp_extra_info['activesheet'] = data.get('activesheet')
        sql_obj.insert_template(template_id, template_name, industry , project, ttype, user, json.dumps(temp_extra_info))
        sheets = []
        rows = []
        sheet_id = 1
        for sheet_info in data['sheets']:
             sheet = self.clean_name(sheet_info['name'])  
             extra_info = copy.deepcopy(sheet_info)
             del extra_info['rows']
             if sheet_info['name'] == data['activeSheet']:
                extra_info['activesheet'] = 1
             sheets.append((template_id, template_name, sheet_id, sheet, user, json.dumps(extra_info)))
             for row in sheet_info['rows']:
                  for colinfo in row['cells']:
                      col = colinfo['index']
                      fromular_str = colinfo.get('formula', '') #self.get_formula_str(col, colinfo)
                      cell_alph = self.find_alp(col)
                      level_id = colinfo.get('level', '')
                      rows.append((template_id, sheet_id, row['index'], col, str(colinfo), 'taxonomy', str(fromular_str), cell_alph, level_id))
             sheet_id = sheet_id + 1
        sql_obj = sqlite_api.sqlite_api(db_path)
        sql_obj.insert_sheets(sheets)
        sql_obj.insert_sheet_data(rows)
        return [{'message': 'done'}]
        

    def insert_sheet_old_template(self, ijson):
        if ijson.get('nw', ''):
            return self.insert_new_template(ijson)
        template_id = ijson['template_id']
        user = ijson['user']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheet_id = sql_obj.max_sheet_id(template_id)
        rows = []
        data = ijson['data']
        sheets = []
        for sheet_info in data['sheets']:
             sheet = self.clean_name(sheet_info['name'])  
             extra_info = copy.deepcopy(sheet_info)
             del extra_info['rows']
             sheets.append((template_id, '', sheet_id, sheet, user, json.dumps(extra_info)))
             for row in sheet_info['rows']:
                  for colinfo in row['cells']:
                      col = colinfo['index']
                      fromular_str = colinfo.get('formula', '') #self.get_formula_str(col, colinfo)
                      cell_alph = self.find_alp(col)
                      level_id = colinfo.get('level', '')
                      rows.append((template_id, sheet_id, row['index'], col, str(colinfo), 'taxonomy', str(fromular_str), cell_alph, level_id))
             sheet_id = sheet_id + 1
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sql_obj = sqlite_api.sqlite_api(db_path)
        sql_obj.insert_sheets(sheets)
        sql_obj.insert_sheet_data(rows)
        return [{'message': 'done'}]
        
 
    def convert_dp_format_templates(self, templates):
        avail_template  = []
        for template in templates:
            avail_template.append({'k': template[0], 'n': template[1]}) 
        return  avail_template 
        #[{'message': 'done', 'drop_down_data': avail_template}]

    def convert_grid_format_templates(self, templates):
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}, {"k":"Template", "n":"Template Name"}], 'data':[], 'map':{}}
        for rid, template in enumerate(templates, 1):
            grid_data['data'].append({'sn': {"v": rid}, 'rid': rid, 'cid': rid, "Template": {"v": template}})
            map_key = "%s_%s"%(rid,"Template")
            grid_data['map'][map_key]  = {} 
        return [grid_data] #[{'message': 'done', 'grid_data': [grid_data]}]
 
    def read_templates(self, ijson):
        project  = ijson['project']
        ttype = ijson.get('type', '')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        if ttype:
            templates = sql_obj.read_templates_type(project, ttype)
        else:
            templates = sql_obj.read_templates(project) 
        res = {"message": "done"}
        res['data'] = self.convert_dp_format_templates(templates)
        return [res]

    def convert_dp_format_sheets(self, sheets):
        avail_sheets  = []
        for sheet in sheets:
            avail_sheets.append({'k': sheet[0], 'n':sheet[1]}) 
        return avail_sheets #[{'message': 'done', 'drop_down_data': avail_sheets}]
        
    def convert_grid_format_sheets(self, sheets):
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}, {"k":"Sheet", "n":"Sheet Name"}], 'data':[], 'map':{}}
        for rid, sheet in enumerate(sheets, 1):
            grid_data['data'].append({'sn': {"v": rid}, 'rid': rid, 'cid': rid, "Sheet": {"v": sheet}})
            map_key = "%s_%s"%(rid,"Template")
            grid_data['map'][map_key]  = {} 
        return grid_data #[{'message': 'done', 'grid_data': [grid_data]}]
        
          
    def read_sheets(self, ijson):
        template_id = ijson['template_id']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheets = sql_obj.read_sheets(template_id) 
        res = {"message": "done"}
        res['drop_down_data'] = self.convert_dp_format_sheets(sheets)
        return [res] 

    def find_alp(self, dec_val):
        if(dec_val<26):
            return chr(65+dec_val)
        if(dec_val > 25 and dec_val < 52):
            return chr(65)+chr(65+(dec_val-26))
        if(dec_val > 51 and dec_val < 78):
            return chr(66)+chr(65+(dec_val-52))
        if(dec_val > 77 and dec_val < 104):
            return chr(67)+chr(65+(dec_val-78))
        return "@@"

    def convert_grid_format_sheetinfo(self, sheet_info):
        fs = {}
        col_def = []
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}], 'grid_data':[], 'map':{}}
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = eval(value)
             fs[row][col]['taxo'] = taxonomy
             if col not in col_def:
                col_def.append(col) 
                grid_data['col_def'].append({'k': col, 'n': cell_alph, 'mw':150, "hcls": "xl_hdr"})          
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = {'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 rr[col] = {'v': colinfo['v']}
                 grid_data['map'][str_cid+"_"+str(col)] = {'taxo': colinfo['taxo']}
              grid_data['grid_data'].append(rr)
         
        return [grid_data]

    def getcellval(self, cellinfo):
        vcel = eval(cellinfo.get('v', {'value':''}))
        if vcel.get('formula', {}):
            if isinstance(vcel.get('value', ''), dict):
                return [vcel['value'].get('code', ''), vcel.get('formula', {}).get('src', ''), vcel]
            else:
                return [vcel['value'], vcel.get('formula', {}).get('src', ''), vcel]
        return [vcel.get('value', ''), '',  vcel]

    def check_column(self, sheet_info):
        col_id = 'not'
        level_id_index = 'not'
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if row != 0:continue
             cdata = eval(value)
             if cdata.get('value', '') == config_obj.taxonomy_column:
                col_id = col
             if cdata.get('value', '') == config_obj.taxonomy_level:
                level_id_index = col
        return col_id, level_id_index
                
        

    def convert_spreadsheet_format_sheetinfo(self, sheet_info, extra_info, ttype, sheet_name):
        fs = {}
        col_def = []
        grid_data = {'columns':[], 'data':[], 'map':{}, "style":{}, 'nested_headers':{}}
        col_info = {}
        for ccol in extra_info.get('columns', []):
            col_info[ccol['index']] = {'w': int(ccol.get('width', 100))}
        col_id = 'not'
        level_id_col = 'not'
        if ttype == 'fixed' and sheet_name == config_obj.taxonomy_sheet:
            col_id, level_id_col = self.check_column(sheet_info)
        #print 'ttyy', col_id
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             print '[[]]', [col_id, level_id_col,]
             if col_id != 'not' and col == level_id_col:
                if row != 0:
                    org_val =  eval(value)
                    grid_data['map'].setdefault('rowGroup', {})
                    grid_data['map']['rowGroup'][row] = len(str(org_val.get('value', '')).split('.'))
             if 0 and col_id != 'not' and col != col_id:continue
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = value
             fs[row][col]['taxo'] = taxonomy
             fs[row][col]['al'] = '%s%s'%(cell_alph, row)
             fs[row][col]['alp'] = cell_alph
             if col not in col_def:
                col_def.append(col) 
                grid_data['columns'].append({'width':col_info.get(col,{}).get('w', 100), 'title': cell_alph, 'type':"text"})
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = [] #{'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 cell_val  = self.getcellval(colinfo)
                 if cell_val[1]:
                    rr.append(cell_val[1]) 
                 else:
                    rr.append(cell_val[0]) 
                 grid_data['map'][colinfo['al']] = {'taxo': colinfo['taxo'], 'r': cid, 'c': col}
                 grid_data['style'][colinfo['al']] = cell_val[2]
              grid_data['data'].append(rr)
        #grid_data['map'] =  {"Restated":{'AK5':5, "AK6": 6, "BB6":10}}     
        return grid_data

    def get_extra_info(self, template_id, sheet_id, sql_obj):
        extra_info = sql_obj.read_sheet_extra_info(template_id, sheet_id)
        return [extra_info[0], json.loads(extra_info[1])]
        
    def convert_table_format_sheetinfo(self, sheet_info):
        fs = {}
        grid_data = {'col_def':[{'k':'sn', 'n':'', 'type':'SL', 'pin':'pinnedLeft', "cls": "xl_hdr", "hcls": "xl_hdr"}], 'table_data':[], 'map':{}}
        col_def = []
        for rowinfo in sheet_info:
             row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
             if not value:
                 value = ''
             fs.setdefault(row, {}).setdefault(col, {})['v'] = value
             fs[row][col]['taxo'] = taxonomy
             fs[row][col]['cell_alph'] = cell_alph
        vrows = fs.keys() 
        vrows.sort()
        for cid , row in enumerate(vrows, 1):
              rinfo = fs[row]
              cols = rinfo.keys()
              cols.sort()
              str_cid = str(cid)
              rr = {'rid': cid, 'cid': cid , 'sn': {'v': cid}}
              for col in cols:
                 colinfo = rinfo[col]
                 if col not in col_def:
                    col_def.append(col) 
                    grid_data['col_def'].append({'k': col, 'n': colinfo["cell_alph"]})          
                 rr[col] = {'v': colinfo['v']}
                 grid_data['map'][str_cid+"_"+str(col)] = {'taxo': colinfo['taxo']}
              grid_data['table_data'].append(rr)
        return grid_data

    def colnum_number(self, col):
        import string
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num


    def getrowspancols(self, s, e):
        sr = int(''.join(re.findall(r'\d+', s)))
        sa = "".join(re.findall("[a-zA-Z]+",s))
        sc = self.colnum_number(sa)
        er = int(''.join(re.findall(r'\d+', e)))
        ea = "".join(re.findall("[a-zA-Z]+", e))
        ec = self.colnum_number(ea)
        return [(er - sr)+1, (ec - sc)+1]
        
 
    def convert_merge_cells(self, mc_lst):
        res = {}
        for cord  in mc_lst:  
            start, end = cord.split(':')
            res[start] = self.getrowspancols(start, end)
            #[rc_info.get('columns', 1), rc_info.get('rows', 1)] 
        return res

    def check_avail_name(self, ijson):
        project = ijson.get('project', '')
        new_name = ijson['name']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        templates = sql_obj.get_avail_name(project, new_name) 
        if templates:
            return [{'message':'done', 'avail':'N'}]
        else:
            return [{'message':'done', 'avail':'Y'}]


    def check_avail_sname(self, ijson):
        template_id = ijson['template_id']
        project = ijson.get('project', '')
        new_name = ijson['name']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        templates = sql_obj.get_avail_sname(project, new_name) 
        if templates:
            return [{'message':'done', 'avail':'N'}]
        else:
            return [{'message':'done', 'avail':'Y'}]
        
        
        
        

    def sheet_gridinfo(self, ijson):
        template_id = ijson['template_id']
        sheet_id   = ijson['sheet_id']
        ttype  = ijson.get('type', 'fixed')
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheet_info = sql_obj.read_sheet_data(template_id, sheet_id) 
        sheet_name, extra_info = self.get_extra_info(template_id, sheet_id, sql_obj)
        res = {"message": "done"}
        res['data'] = self.convert_spreadsheet_format_sheetinfo(sheet_info, extra_info, ttype, sheet_name)
        if extra_info.get('mergedCells', []):
            res['data']['mergeCells'] = self.convert_merge_cells(extra_info['mergedCells'])
        return [res] 
        

    def create_excel(self, ijson):
        template_id = ijson['template_id']
        import sqlite_api
        db_path = config_obj.Template_storage
        sql_obj = sqlite_api.sqlite_api(db_path)
        sheets = sql_obj.read_sheets(template_id) 
        res = {}
        property_dic = {}
        order_sheets = []
        for sheet in sheets:
            idd, sheet_name = sheet
            sheet_info = sql_obj.read_sheet_data(template_id, idd) 
            sheet_name, extra_info = self.get_extra_info(template_id, idd, sql_obj)
            order_sheets.append(idd)
            res.setdefault(idd, {})['name'] = sheet_name
            number_rows = 0
            number_cols = 0
            sheet_info = sql_obj.read_sheet_data(template_id, idd) 
            for rowinfo in sheet_info:
                row, col, value, taxonomy, fromular_str, cell_alph = rowinfo 
                if row >number_rows:    
                    number_rows = row
                if col > number_cols:
                    number_cols =  col
        
                temp = {}
                org_val= eval(value)
                if org_val.get('formula', {}):
                    temp['excel_formula'] = org_val['formula']['src']
                    if isinstance(org_val['value'],  dict):
                        temp['value'] = org_val['value']['code']
                    else:
                        temp['value'] = org_val['value']
                else:
                    temp['value'] =  org_val.get('value', '')
                res[idd].setdefault('cell_dict', {})["%s_%s"%(row, col)] = temp
                #res[idd].setdefault('cell_dict', {})[(row, col)] = temp
                for kk,vv in org_val.items():
                    property_dic[kk] = vv
            res[idd]['nrows'] = number_rows + 1
            res[idd]['ncols'] = number_cols + 1
            res[idd]['merge_range'] = self.get_merged_range(extra_info['mergedCells'])
                
        import create_excel as ce
        ce_obj = ce.xlsx_operation()
        order_sheets.sort()
        import datetime 
        cp = datetime.datetime.now().strftime("%m%d%Y%H%M%S")
        path = "/var/www/html/sudhakar/%s.xlsx"%(cp)
        ce_obj.writeExcel(path, res, order_sheets)
        return [{'message':'done', 'path': path, "data": res, 'pro': property_dic}]
        
        
    def get_merged_range(self, mc_lst):
        res = []
        for cord  in mc_lst:  
            s, e = cord.split(':')
            sr = int(''.join(re.findall(r'\d+', s)))
            sa = "".join(re.findall("[a-zA-Z]+",s))
            er = int(''.join(re.findall(r'\d+', e)))
            ea = "".join(re.findall("[a-zA-Z]+", e))
            res.append([sr, s, er, e])
        return res
                

    def read_only_taxo(self, ijson):
        project_id = ijson['project']
        template_id = ijson['template_id']
        db_path    = config.Config.taxonomy_path.format(project_id, template_id)+"info.db"
        import sqlite_api
        sql_obj = sqlite_api.sqlite_api(db_path)
        res = sql_obj.get_taxo_and_level()
        grid_data = {'columns':[{"width": 159, "type": "text", "title": "A"}], 'data':[], 'map':{}, "style":{}, 'nested_headers':{}}
        for row, dd in enumerate(res):
            taxo, level = dd  
            grid_data['map'].setdefault('rowGroup', {})
            grid_data['map']['rowGroup'][row] = level
            grid_data['data'].append([taxo])
        return [{'message':'done', 'data': grid_data}]
                

if __name__ == '__main__': 
    obj = model_api()
    ijson = {"path":{"filename": "http:/172.16.20.229/muthu/test_excel/GA_Merge.xlsx"}, "user":"user", "project":"INC", "industry": "2", "template_id": 1,'sheet_id':1, "wf":[{"mod":"grid"}]}
    #print json.dumps(obj.insert_model(ijson))
    #print json.dumps(obj.read_templates(ijson))
    #print json.dumps(obj.read_sheets(ijson))
    #print json.dumps(obj.sheet_gridinfo(ijson))
